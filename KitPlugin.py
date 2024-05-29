import csv
import math
import os

import openpyxl
import pandas as pd

from Kitsune import Kitsune
from KitNET.KitNET import KitNET
import netStat as ns
import shap
import numpy as np
import pickle
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import sklearn
import optuna
from scipy.stats import norm
import random
from scapy.all import PcapReader, PcapWriter, wrpcap, rdpcap, IP, TCP, UDP

# Class that provides a callable interface for Kitsune components.
# Guy Puts, 2024
class KitPlugin:
    # Function used by SHAP as callback to test instances of features
    def kitsune_model(self, input_data):
        prediction = self.K.feed_batch(input_data)
        return prediction

    def kitnet_model(self, input_data):
        prediction = self.KitTest.process_batch(input_data)
        return prediction

    # Builds a Kitsune instance. Does not train KitNET yet.
    def __init__(self, input_path=None, packet_limit=None, num_autenc=None, FMgrace=None, ADgrace=None, learning_rate=0.1, hidden_ratio=0.75, num_features=420):
        # This code will be removed when batch running Kitsune has been finalized
        if input_path != None and num_autenc != None:
            self.features_list = None
            self.explainer = None
            self.shap_values = None
            self.K = Kitsune(input_path, packet_limit, num_autenc, FMgrace, ADgrace, learning_rate, hidden_ratio)
            self.metadata = {
                "filename" : input_path,
                "packet_limit" : packet_limit,
                "num_autenc" : num_autenc,
                "FMgrace": FMgrace,
                "ADgrace" : ADgrace,
                "timestamp" : datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
        self.testFeatures = None
        maxHost = 100000000000
        maxSess = 100000000000
        self.nstat = ns.netStat(np.nan, maxHost, maxSess)
        self.num_features = num_features

    # Calls Kitsune's get_feature_list function to build the list of features
    def feature_builder(self, csv=False, single=False, kind=1):
        print("Building features")
        # Dummy-running Kitsune to get a list of features
        self.features_list = self.K.get_feature_list(csv, single, kind=kind)
        return self.features_list

    # Loads Kitsune's feature list from a pickle file
    def feature_loader(self, newpickle=None):
        print("Loading features from file")
        path = 'pickles/featureList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'rb') as f:
            features_list = pickle.load(f)
        self.features_list = features_list

    # Writes Kitsune's feature list to a pickle file
    def feature_pickle(self, newpickle=None):
        print("Writing features to file")
        path = 'pickles/featureList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'wb') as f:
            pickle.dump(self.features_list, f)

    # Trains KitNET, using the specified index range of this class' feature list
    def kit_trainer(self, min_index, max_index):
        print("Training")
        self.K.feed_batch(self.features_list[min_index:max_index])
        print("Training finished")

    # Trains KitNET, using a supplied feature list
    def kit_trainer_supplied_features(self, features_list):
        print("Training")
        self.K.feed_batch(features_list)
        print("Training finished")

    # Runs KitNET, using specified index range of this class' feature list
    def kit_runner(self, min_index, max_index, normalize=False):
        print("Running")
        print(len(self.features_list[min_index:max_index]))
        return self.K.feed_batch(self.features_list[min_index:max_index])

    # Calculates the three best and worst values for all statistics
    def get_high_low_indices(self):
        shap_transposed = self.shap_values.T
        # List of statistics functions
        stat_functions = {
            'mean': np.mean,
            'median': np.median,
            'std_dev': np.std,
            'variance': np.var,
            'minimum': np.min,
            'maximum': np.max,
            'total_sum': np.sum
        }

        # Dictionary to store results
        result_dict = {}

        # Loop over statistics
        for stat_name, stat_func in stat_functions.items():
            # Calculate the statistic for each list
            stat_values = stat_func(shap_transposed, axis=1)

            # Calculate the indices of the highest and lowest values
            sorted_indices = np.argsort(stat_values)
            highest_indices = sorted_indices[-3:]
            lowest_indices = sorted_indices[:3]
            # Store the indices in the result dictionary
            result_dict[stat_name] = {
                'highest_indices': highest_indices,
                'lowest_indices': lowest_indices
            }
        return result_dict

    # Creates an Excel sheet with relevant statistics
    def create_sheet(self, day, sheet_title):
        sheet = self.workbook.copy_worksheet(self.workbook.active)
        sheet.title = sheet_title
        headers = ['Mean', 'Median', 'Standard Deviation', 'Variance', 'Minimum', 'Maximum', 'Sum', 'Metadata']
        header_row = headers
        lambdameans = {
            '5':[],
            '3':[],
            '1':[],
            '0.1':[],
            '0.01':[]
        }
        lambdamean = {
            '5': None,
            '3': None,
            '1': None,
            '0.1': None,
            '0.01': None
        }
        lambdamedians = {
            '5':[],
            '3':[],
            '1':[],
            '0.1':[],
            '0.01':[]
        }
        lambdamedian = {
            '5': None,
            '3': None,
            '1': None,
            '0.1': None,
            '0.01': None
        }
        featuremeans = {
            'weight':[],
            'mean':[],
            'variance':[],
            'radius':[],
            'magnitude':[],
            'covariance':[],
            'pearson correlation coefficient':[],
            'weight (JIT)':[],
            'mean (JIT)':[],
            'variance (JIT)':[],
            'median (JIT)':[],
            'TCP FIN frequency':[],
            'TCP SYN frequency':[],
            'TCP RST frequency':[],
            'TCP PSH frequency':[],
            'TCP ACK frequency':[],
            'TCP URG frequency':[],
            'TCP ECE frequency':[],
            'TCP CWR frequency':[],
            'TCP Flag count':[],
            '25th Quantile datagram size':[],
            '50th Quantile datagram size (median)':[],
            '75th Quantile datagram size':[]
        }
        featuremean = {
            'weight':None,
            'mean':None,
            'variance':None,
            'radius':None,
            'magnitude':None,
            'covariance':None,
            'pearson correlation coefficient':None,
            'weight (JIT)':None,
            'mean (JIT)':None,
            'variance (JIT)':None,
            'median (JIT)':None,
            'TCP FIN frequency':None,
            'TCP SYN frequency':None,
            'TCP RST frequency':None,
            'TCP PSH frequency':None,
            'TCP ACK frequency':None,
            'TCP URG frequency':None,
            'TCP ECE frequency':None,
            'TCP CWR frequency':None,
            'TCP Flag count':None,
            '25th Quantile datagram size':None,
            '50th Quantile datagram size (median)':None,
            '75th Quantile datagram size':None
        }
        featuremedians = {
            'weight':[],
            'mean':[],
            'variance':[],
            'radius':[],
            'magnitude':[],
            'covariance':[],
            'pearson correlation coefficient':[],
            'weight (JIT)':[],
            'mean (JIT)':[],
            'variance (JIT)':[],
            'median (JIT)':[],
            'TCP FIN frequency':[],
            'TCP SYN frequency':[],
            'TCP RST frequency':[],
            'TCP PSH frequency':[],
            'TCP ACK frequency':[],
            'TCP URG frequency':[],
            'TCP ECE frequency':[],
            'TCP CWR frequency':[],
            'TCP Flag count':[],
            '25th Quantile datagram size':[],
            '50th Quantile datagram size (median)':[],
            '75th Quantile datagram size':[]
        }
        featuremedian = {
            'weight':None,
            'mean':None,
            'variance':None,
            'radius':None,
            'magnitude':None,
            'covariance':None,
            'pearson correlation coefficient':None,
            'weight (JIT)':None,
            'mean (JIT)':None,
            'variance (JIT)':None,
            'median (JIT)':None,
            'TCP FIN frequency':None,
            'TCP SYN frequency':None,
            'TCP RST frequency':None,
            'TCP PSH frequency':None,
            'TCP ACK frequency':None,
            'TCP URG frequency':None,
            'TCP ECE frequency':None,
            'TCP CWR frequency':None,
            'TCP Flag count':None,
            '25th Quantile datagram size':None,
            '50th Quantile datagram size (median)':None,
            '75th Quantile datagram size':None
        }

        aggtypemeans = {
            'SRC IP':[],
            'Channel':[],
            'Socket':[],
            'DST IP':[],
            'SRC IP (JIT)':[],
            'Channel (JIT)':[],
            'Socket (JIT)':[],
            'DST IP (JIT)':[]

        }
        aggtypemean = {
            'SRC IP': None,
            'Channel': None,
            'Socket': None,
            'DST IP': None,
            'SRC IP (JIT)': None,
            'Channel (JIT)': None,
            'Socket (JIT)': None,
            'DST IP (JIT)': None
        }
        aggtypemedians = {
            'SRC IP': [],
            'Channel': [],
            'Socket': [],
            'DST IP': [],
            'SRC IP (JIT)': [],
            'Channel (JIT)': [],
            'Socket (JIT)': [],
            'DST IP (JIT)': []
        }
        aggtypemedian = {
            'SRC IP': None,
            'Channel': None,
            'Socket': None,
            'DST IP': None,
            'SRC IP (JIT)': None,
            'Channel (JIT)': None,
            'Socket (JIT)': None,
            'DST IP (JIT)': None
        }

        for col, value in enumerate(header_row):
            cell = sheet.cell(row=1, column=6 + col)
            cell.value = value
        for idx, num_list in enumerate(self.shap_values.T):
            #num_list = abs(num_list)
            mean = np.mean(num_list)
            median = np.median(num_list)
            std_dev = np.std(num_list)
            variance = np.var(num_list)
            minimum = np.min(num_list)
            maximum = np.max(num_list)
            total_sum = np.sum(num_list)
            if idx+1 in [1,2,3,16,17,18,19,20,21,22,51,52,53,54,71,72,73,74,75,76,77,106,107,108,109,126,127,128,129,146,147,148,161,162,163,164,181,182,183,184,185,186,187,188,189,226,227,228,229,230,231,232,233,234,271,272,273,274,275,276,277,278,279,316,317,318,319,320,321,322,323,324,361,362,363,376,377,378,391,392,393,406,407,408]:
                lambdameans['5'].append(mean)
                lambdamedians['5'].append(median)
            if idx+1 in [4,5,6,23,24,25,26,27,28,29,55,56,57,58,78,79,80,81,82,83,84,110,111,112,113,130,131,132,133,149,150,151,165,166,167,168,190,191,192,193,194,195,196,197,198,235,236,237,238,239,240,241,242,243,280,281,282,283,284,285,286,287,288,325,326,327,328,329,330,331,332,333,364,365,366,379,380,381,394,395,396,409,410,411]:
                lambdameans['3'].append(mean)
                lambdamedians['3'].append(median)
            if idx+1 in [7,8,9,30,31,32,33,34,35,36,59,60,61,62,85,86,87,88,89,90,91,114,115,116,117,134,135,136,137,152,153,154,169,170,171,172,199,200,201,202,203,204,205,206,207,244,245,246,247,248,249,250,251,252,289,290,291,292,293,294,295,296,297,334,335,336,337,338,339,340,341,342,367,368,369,382,383,384,397,398,399,412,413,414]:
                lambdameans['1'].append(mean)
                lambdamedians['1'].append(median)
            if idx+1 in [10,11,12,37,38,39,40,41,42,43,63,64,65,66,92,93,94,95,96,97,98,118,119,120,121,138,139,140,141,155,156,157,173,174,175,176,208,209,210,211,212,213,214,215,216,253,254,255,256,257,258,259,260,261,298,299,300,301,302,303,304,305,306,343,344,345,346,347,348,349,350,351,370,371,372,385,386,387,400,401,402,415,416,417]:
                lambdameans['0.1'].append(mean)
                lambdamedians['0.1'].append(median)
            if idx+1 in [13,14,15,44,45,46,47,48,49,50,67,68,69,70,99,100,101,102,103,104,105,122,123,124,125,142,143,144,145,158,159,160,177,178,179,180,217,218,219,220,221,222,223,224,225,262,263,264,265,266,267,268,269,270,307,308,309,310,311,312,313,314,315,352,353,354,355,356,357,358,359,360,373,374,375,388,389,390,403,404,405,418,419,420]:
                lambdameans['0.01'].append(mean)
                lambdamedians['0.01'].append(median)

            if idx+1 in [1,4,7,10,13,16,23,30,37,44,71,78,85,92,99,146,149,152,155,158]:
                featuremeans['weight'].append(mean)
                featuremedians['weight'].append(median)
            if idx+1 in [2,5,8,11,14,17,24,31,38,45,72,79,86,93,100,147,150,153,156,159]:
                featuremeans['mean'].append(mean)
                featuremedians['mean'].append(median)
            if idx + 1 in [3,6,9,12,15,18,25,32,39,46,73,80,87,94,101,148,151,154,157,160]:
                featuremeans['variance'].append(mean)
                featuremedians['variance'].append(median)
            if idx + 1 in [19,26,33,40,47,74,81,88,95,102]:
                featuremeans['radius'].append(mean)
                featuremedians['radius'].append(median)
            if idx + 1 in [20,27,34,41,48,75,82,89,96,103]:
                featuremeans['magnitude'].append(mean)
                featuremedians['magnitude'].append(median)
            if idx + 1 in [21,28,35,42,49,76,83,90,97,104]:
                featuremeans['covariance'].append(mean)
                featuremedians['covariance'].append(median)
            if idx + 1 in [22,29,36,43,50,77,84,91,98,105]:
                featuremeans['pearson correlation coefficient'].append(mean)
                featuremedians['pearson correlation coefficient'].append(median)
            if idx + 1 in [51,55,59,63,67,106,110,114,118,122,126,130,134,138,142,161,165,169,173,177]:
                featuremeans['weight (JIT)'].append(mean)
                featuremedians['weight (JIT)'].append(median)
            if idx + 1 in [52,56,60,64,68,107,111,115,119,123,127,131,135,139,143,162,166,170,174,178]:
                featuremeans['mean (JIT)'].append(mean)
                featuremedians['mean (JIT)'].append(median)
            if idx + 1 in [53,57,61,65,69,108,112,116,120,124,128,132,136,140,144,163,167,171,175,179]:
                featuremeans['variance (JIT)'].append(mean)
                featuremedians['variance (JIT)'].append(median)
            if idx + 1 in [54,58,62,66,70,109,113,117,121,125,129,133,137,141,145,164,168,172,176,180]:
                featuremeans['median (JIT)'].append(mean)
                featuremedians['median (JIT)'].append(median)
            if idx + 1 in [181,190,199,208,217,226,235,244,253,262,271,280,289,298,307,316,325,334,343,352]:
                featuremeans['TCP FIN frequency'].append(mean)
                featuremedians['TCP FIN frequency'].append(median)
            if idx + 1 in [182,191,200,209,218,227,236,245,254,263,272,281,290,299,308,317,326,335,344,353]:
                featuremeans['TCP SYN frequency'].append(mean)
                featuremedians['TCP SYN frequency'].append(median)
            if idx + 1 in [183,192,201,210,219,228,237,246,255,264,273,282,291,300,309,318,327,336,345,354]:
                featuremeans['TCP RST frequency'].append(mean)
                featuremedians['TCP RST frequency'].append(median)
            if idx + 1 in [184,193,202,211,220,229,238,247,256,265,274,283,292,301,310,319,328,337,346,355]:
                featuremeans['TCP PSH frequency'].append(mean)
                featuremedians['TCP PSH frequency'].append(median)
            if idx + 1 in [185,194,203,212,221,230,239,248,257,266,275,284,293,302,311,320,329,338,347,356]:
                featuremeans['TCP ACK frequency'].append(mean)
                featuremedians['TCP ACK frequency'].append(median)
            if idx + 1 in [186,195,204,213,222,231,240,249,258,267,276,285,294,303,312,321,330,339,348,357]:
                featuremeans['TCP URG frequency'].append(mean)
                featuremedians['TCP URG frequency'].append(median)
            if idx + 1 in [187,196,205,214,223,232,241,250,259,268,277,286,295,304,313,322,331,340,349,358]:
                featuremeans['TCP ECE frequency'].append(mean)
                featuremedians['TCP ECE frequency'].append(median)
            if idx + 1 in [188,197,206,215,224,233,242,251,260,269,278,287,296,305,314,323,332,341,350,359]:
                featuremeans['TCP CWR frequency'].append(mean)
                featuremedians['TCP CWR frequency'].append(median)
            if idx + 1 in [189,198,207,216,225,234,243,252,261,270,279,288,297,306,315,324,333,342,351,360]:
                featuremeans['TCP Flag count'].append(mean)
                featuremedians['TCP Flag count'].append(median)
            if idx + 1 in [361,364,367,370,373,376,379,382,385,388,391,394,397,400,403,406,409,412,415,418]:
                featuremeans['25th Quantile datagram size'].append(mean)
                featuremedians['25th Quantile datagram size'].append(median)
            if idx + 1 in [362,365,368,371,374,377,380,383,386,389,392,395,398,401,404,407,410,413,416,419]:
                featuremeans['50th Quantile datagram size (median)'].append(mean)
                featuremedians['50th Quantile datagram size (median)'].append(median)
            if idx + 1 in [363,366,369,372,375,378,381,384,387,390,393,396,399,402,405,408,411,414,417,420]:
                featuremeans['75th Quantile datagram size'].append(mean)
                featuremedians['75th Quantile datagram size'].append(median)

            if idx+1 in [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,210,211,212,213,214,215,216,217,218,219,220,221,222,223,224,225,361,362,363,364,365,366,367,368,369,370,371,372,373,374,375]:
                aggtypemeans['SRC IP'].append(mean)
                aggtypemedians['SRC IP'].append(median)
            if idx+1 in [16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,226,227,228,229,230,231,232,233,234,235,236,237,238,239,240,241,242,243,244,245,246,247,248,249,250,251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,376,377,378,379,380,381,382,383,384,385,386,387,388,389,390]:
                aggtypemeans['Channel'].append(mean)
                aggtypemedians['Channel'].append(median)
            if idx+1 in [71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,300,301,302,303,304,305,306,307,308,309,310,311,312,313,314,315,391,392,393,394,395,396,397,398,399,400,401,402,403,404,405]:
                aggtypemeans['Socket'].append(mean)
                aggtypemedians['Socket'].append(median)
            if idx+1 in [146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,316,317,318,319,320,321,322,323,324,325,326,327,328,329,330,331,332,333,334,335,336,337,338,339,340,341,342,343,344,345,346,347,348,349,350,351,352,353,354,355,356,357,358,359,360,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420]:
                aggtypemeans['DST IP'].append(mean)
                aggtypemedians['DST IP'].append(median)
            if idx+1 in [106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125]:
                aggtypemeans['SRC IP (JIT)'].append(mean)
                aggtypemedians['SRC IP (JIT)'].append(median)
            if idx+1 in [51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70]:
                aggtypemeans['Channel (JIT)'].append(mean)
                aggtypemedians['Channel (JIT)'].append(median)
            if idx+1 in [126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145]:
                aggtypemeans['Socket (JIT)'].append(mean)
                aggtypemedians['Socket (JIT)'].append(median)
            if idx+1 in [161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180]:
                aggtypemeans['DST IP (JIT)'].append(mean)
                aggtypemedians['DST IP (JIT)'].append(median)
            row_data = [mean, median, std_dev, variance, minimum, maximum, total_sum]

            for col, value in enumerate(row_data):
                cell = sheet.cell(row=idx + 2, column=6 + col)
                cell.value = value

        row = idx + 2
        row += 1
        cell = sheet.cell(row=row, column=1)
        cell.value = "Grouped by lambda"
        cell = sheet.cell(row=row, column=2)
        cell.value = "mean"
        cell = sheet.cell(row=row, column=3)
        cell.value = "median"
        row += 1

        for key in lambdameans:
            cell = sheet.cell(row=row, column = 1)
            cell.value = key
            cell = sheet.cell(row=row, column = 2)
            cell.value = np.mean(np.array(lambdameans[key]))
            lambdamean[key] = np.mean(np.array(lambdameans[key]))
            cell = sheet.cell(row=row, column=3)
            cell.value = np.median(np.array(lambdamedians[key]))

            lambdamedian[key] = np.median(np.array(lambdamedians[key]))
            row += 1
        row += 1

        cell = sheet.cell(row=row, column=1)
        cell.value = "Grouped by feature"
        cell = sheet.cell(row=row, column=2)
        cell.value = "mean"
        cell = sheet.cell(row=row, column=3)
        cell.value = "median"
        row += 1
        for key in featuremeans:
            cell = sheet.cell(row=row, column=1)
            cell.value = key
            cell = sheet.cell(row=row, column=2)
            cell.value = np.mean(np.array(featuremeans[key]))
            featuremean[key] = np.mean(np.array(featuremeans[key]))
            cell = sheet.cell(row=row, column=3)
            cell.value = np.median(np.array(featuremedians[key]))
            print(f"key: {key}, value: {np.mean(np.array(featuremeans[key]))}")
            featuremedian[key] = np.median(np.array(featuremedians[key]))
            row += 1

        cell = sheet.cell(row=row, column=1)
        cell.value = "Grouped by aggregation type"
        cell = sheet.cell(row=row, column=2)
        cell.value = "mean"
        cell = sheet.cell(row=row, column=3)
        cell.value = "median"
        row += 1
        for key in aggtypemeans:
            cell = sheet.cell(row=row, column=1)
            cell.value = key
            cell = sheet.cell(row=row, column=2)
            cell.value = np.mean(np.array(aggtypemeans[key]))
            aggtypemean[key] = np.mean(np.array(aggtypemeans[key]))
            cell = sheet.cell(row=row, column=3)
            cell.value = np.median(np.array(aggtypemedians[key]))
            print(f"key: {key}, value: {np.mean(np.array(aggtypemeans[key]))}")
            aggtypemedian[key] = np.median(np.array(aggtypemedians[key]))
            row += 1

        # self.create_histogram(day, featuremean, featuremedian, sheet_title+" grouped by feature name")
        # self.create_histogram_to_sheet_feature(day, featuremean, featuremedian, sheet_title + " grouped by feature name", sheet, "A")
        # self.create_histogram(day, lambdamean, lambdamedian, sheet_title + " grouped by lambda value")
        # self.create_histogram_to_sheet_lambda(day, lambdamean, lambdamedian, sheet_title + " grouped by lambda value", sheet, "F")
        row += 1

        color_indices = self.get_high_low_indices()
        stat_columns = {
            'mean': "F",
            'median': "G",
            'std_dev': "H",
            'variance': "I",
            'minimum': "J",
            'maximum': "K",
            'total_sum': "L"
        }
        for stat in color_indices:
            for index in color_indices[stat]["highest_indices"]:
                cell_index = stat_columns[stat] + str(index + 2)
                if stat == "std_dev" or stat == "variance":
                    # Make largest three standard deviation and variance values blue
                    sheet[cell_index].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif stat == "minimum":
                    # Make largest three cells minimum red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    # In all other cases, make largest three cells green
                    sheet[cell_index].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for index in color_indices[stat]["lowest_indices"]:
                cell_index = stat_columns[stat] + str(index + 2)
                if stat == "minimum":
                    # Make largest three cells minimum red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    # In all other cases, make smallest three cells red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # Fill in metadata
        start_row = 2
        start_column_keys = 'M'
        start_column_values = 'N'

        # Loop over the dictionary and write capitalized keys and values to cells
        if hasattr(self, "metadata"):
            for idx, (key, value) in enumerate(self.metadata.items()):
                capitalized_key = key[0].upper() + key[1:]
                key_cell = f"{start_column_keys}{start_row + idx}"
                value_cell = f"{start_column_values}{start_row + idx}"
                sheet[key_cell] = capitalized_key
                sheet[value_cell] = value
        return sheet

    # Calculates an EER-score for a list of RMSEs
    def calc_eer(self, RMSEs, labels):
        fpr, tpr, threshold = sklearn.metrics.roc_curve(labels, RMSEs, pos_label=1)
        fnr = 1-tpr
        #eer_threshold = threshold[np.nanargmin(np.absolute((fnr-fpr)))]
        EER = fpr[np.nanargmin(np.absolute((fnr-fpr)))]
        return EER

    # Calculates an AUC-score for a list of RMSEs and a list of expected values
    def calc_auc(self, RMSEs, labels):
        auc_score = sklearn.metrics.roc_auc_score(labels, RMSEs)
        return auc_score

    # Calculates an EER-score for a list of RMSEs and a list of expected values
    def calc_auc_eer(self, RMSEs, labels):
        return (self.calc_auc(RMSEs, labels), self.calc_eer(RMSEs, labels))

    def read_label_file(self, csvpath):
        with open(csvpath, newline='') as csvfile:
            returnList = []
            labelreader = csv.reader(csvfile, delimiter=' ')
            for row in labelreader:
                row = row[0].strip('][').split(',')
                returnList.append(row)
            return returnList

    def sample_packets_by_conversation(self, tsvpath, outpath, labels):
        # We open the output writer to write to a new TSV file
        with open(outpath, 'w') as op:
            wr = csv.writer(op)
            # We open the reader to get the packets from the original TSV file
            with open(tsvpath) as fd:
                rd = csv.reader(fd, delimiter="\t", quotechar='"')
                pkt_iter = -1
                for row in rd:
                    if pkt_iter % 10000 == 0:
                        print(pkt_iter)
                    if pkt_iter == -1:
                        pkt_iter += 1
                        continue
                    # Labels is the list of conversations, that has previously been sampled to 10 percent of conversations
                    for label in labels:
                        if label[0] == 'Src' or label[0] == 'id' or not label:
                            continue
                        if len(row) < 4:
                            row = row[0].split('\t')
                        if (row[4] == label[0] and row[6] == label[1] and row[5] == label[2] and row[7] == label[3]) or (row[4] == label[2] and row[6] == label[3] and row[5] == label[0] and row[7] == label[1]):
                            print('match')
                            label_iter = label[5]
                            label_val = label[4]
                            row.append(str(pkt_iter))
                            row.append(str(label_iter))
                            row.append(str(label_val))
                            wr.writerow(row)
                            break
                    pkt_iter += 1
            op.close()

    def map_packets_to_features(self, packet_path, feature_path, sampled_feature_path):
        # Step 1: Read the packet TSV file and create a set of packet indices
        subset_indices = set()
        row_index = 0
        print('reading packets')
        with open(packet_path, 'r', newline='') as packet_file:
            csvreader = csv.reader(packet_file)
            for row in csvreader:
                if row:
                    #packet_index = int(row[19])  # Assuming index is in the 20th column
                    packet_index = int(row[-3])  # Assuming index is in the 23rd column
                    subset_indices.add(packet_index)
                row_index += 1
        # Step 2: Read the required statistics from the large feature CSV file
        # and write them to the output CSV file
        print('reading features')
        print(subset_indices)
        with open(feature_path, 'r', newline='') as feature_file, open(sampled_feature_path, 'w', newline='') as output_file:
            csvreader = csv.reader(feature_file)
            csvwriter = csv.writer(output_file)

            counter = 0
            for row_num, row in enumerate(csvreader, start=1):
                packet_index = row_num  # Index is the row number
                # Check if the packet index is in the list of subset indices
                counter += 1
                if packet_index in subset_indices:
                    # Write the row to the output CSV file
                    print('match on feature')
                    csvwriter.writerow(row)

    # Runs a hyperparameter optimization on the supplied dataset, constrained by number of runs and packet limit
    # This version uses KitNET directly instead of running Kitsune as a whole
    def hyper_opt_KitNET(self, day, attack_type, training_cutoff):
        def objective(trial):
            numAE = trial.suggest_int('numAE', 0, 200)
            learning_rate = trial.suggest_float('learning_rate', 0, 0.5)
            hidden_ratio = trial.suggest_float('hidden_ratio', 0, 1)
            FMgrace = trial.suggest_int('FMgrace', 0, 5000000)

            kit = KitNET(self.num_features, max_autoencoder_size=numAE, FM_grace_period=FMgrace, AD_grace_period=math.floor(training_cutoff*0.9), learning_rate=learning_rate, hidden_ratio=hidden_ratio)
            # Load the feature list beforehand to save time
            counter = 0
            if attack_type == "all":
                path = f"input_data/attack_types/{day}_features.csv"
            else:
                path = f"input_data/attack_types/{day}_features_{attack_type}.csv"
            with open(path) as fp:
                rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
                train_err = []
                for packet in rd_ft:
                    if packet:
                        packet = packet[0].split(',')
                        packet = [float(element) for element in packet]
                        packet = np.array(packet)
                        if counter % 10000 == 0:
                            print("training: "+str(counter))
                        train_err.append(kit.train(packet))
                        counter += 1
                    if counter >= training_cutoff:
                        break
                fp.close()
            trial.set_user_attr("train_packets", counter)
            if attack_type == "all":
                conv_train_err = self.map_results_to_conversation(train_err,
                                                                  f"input_data/attack_types/{day}_all.pcap.tsv")
            else:
                conv_train_err = self.map_results_to_conversation(train_err, f"input_data/attack_types/{day}_{attack_type}.pcap.tsv")
            conv_train_err = [max(values) for values in conv_train_err.values()]

            y_pred = []
            path = 'pickles/medium_validate.pkl'
            counterValidate = 0
            print('reading validate list')
            with open(path, 'rb') as f:
                validateList = pickle.load(f)
            for packet in validateList:
                if packet:
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    score = kit.execute(packet)
                    if counterValidate % 10000:
                        print("testing: "+str(counterValidate))
                    y_pred.append(score)
            conv_y_pred = self.map_results_to_conversation(y_pred, f"input_data/attack_types/noday_UNSW_Benign_medium_validate.pcap.tsv")
            conv_y_pred = [max(values) for values in conv_y_pred.values()]
            trial.set_user_attr("training_error", np.mean(conv_train_err))
            trial.set_user_attr("train_median", np.median(conv_train_err))
            trial.set_user_attr("train_25_percentile", np.percentile(conv_train_err, 25))
            trial.set_user_attr("train_75_percentile", np.percentile(conv_train_err, 75))
            trial.set_user_attr("train_max", np.max(conv_train_err))
            trial.set_user_attr("testing_error", np.mean(conv_y_pred))
            trial.set_user_attr("test_median", np.median(conv_y_pred))
            trial.set_user_attr("test_25_percentile", np.percentile(conv_y_pred, 25))
            trial.set_user_attr("test_75_percentile", np.percentile(conv_y_pred, 75))
            trial.set_user_attr("test_max", np.max(conv_y_pred))

            median_value = np.median(conv_train_err)
            median_absolute_deviation = np.median([abs(number - median_value) for number in conv_train_err])
            trial.set_user_attr("mad", median_absolute_deviation)

            threshold = median_value + 2 * median_absolute_deviation
            threshold_one = median_value + median_absolute_deviation
            threshold_median = median_value

            trial.set_user_attr("threshold", threshold)
            trial.set_user_attr("threshold_one", threshold)

            trial.set_user_attr("test_minus_train_error", np.mean(conv_y_pred)-np.mean(conv_train_err))

            anomaly_count = 0
            for err in conv_y_pred:
                if err > threshold:
                    anomaly_count += 1

            trial.set_user_attr("anomaly_count", anomaly_count)
            anomaly_count = 0
            for err in conv_y_pred:
                if err > threshold_one:
                    anomaly_count += 1
            FPR = anomaly_count / len(conv_y_pred)
            trial.set_user_attr("anomaly_count_one", anomaly_count)
            anomaly_count = 0
            for err in conv_y_pred:
                if err > threshold_median:
                    anomaly_count += 1
            trial.set_user_attr("anomaly_count_median", anomaly_count)
            trial.set_user_attr("train_convs", len(train_err))
            trial.set_user_attr("test_convs", len(conv_y_pred))

            return FPR

        # Dashboard logic
        search_space = {
            'numAE': [25, 50, 75],
            'learning_rate': [0.0005, 0.001, 0.005, 0.01, 0.05, 0.1, 0.15],
            'hidden_ratio': [0.25, 0.5, 0.75],
            'FMgrace': [math.floor(0.05*training_cutoff), math.floor(0.10*training_cutoff), math.floor(0.20 * training_cutoff)]
        }
        study = optuna.create_study(sampler=optuna.samplers.GridSampler(search_space), storage=f"sqlite:///{attack_type}.db", study_name=attack_type, load_if_exists=True)
        study.optimize(objective, n_trials=3*9*3*3)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        header = ["Trial Number", "numAE", "learning_rate", "hidden_ratio"]
        ws.append(header)

        # Write trial information
        best_value = float("inf")
        best_row_idx = None  # Track the index of the best row
        for idx, trial in enumerate(study.trials, start=2):  # Start from row 2 to leave room for the header
            trial_params = trial.params
            trial_row = [trial.number, trial_params["numAE"], trial_params["learning_rate"],
                         trial_params["hidden_ratio"], trial.value]
            ws.append(trial_row)

            if trial.value < best_value:
                best_value = trial.value
                best_row_idx = idx

        # Set fill color for the best value row
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        if best_row_idx is not None:
            for cell in ws[best_row_idx]:
                cell.fill = green_fill

        # Save the workbook to a file
        excel_file_path = "output_data/hyperparameter_optimization_results_" + datetime.now().strftime(
            '%d-%m-%Y_%H-%M') + ".xlsx"
        wb.save(excel_file_path)

        print("Results exported to", excel_file_path)
        return study.best_trial

    def run_trained_kitsune_from_feature_csv(self, test_path, test_start, test_limit, kit_path=False):
        if kit_path:
            with open(kit_path, 'rb') as f:
                kit = pickle.load(f)
        else:
            with open("pickles/anomDetectorFullDataset.pkl", 'rb') as f:
                kit = pickle.load(f)

        counter = 0
        results = []
        import csv
        import time

        with open(test_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            start_time = time.time()
            packets_processed = 0

            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    if counter % 10000 == 0:
                        print('running: ')
                        print(counter)
                    if counter < test_limit:
                        results.append(kit.execute(packet))
                        counter += 1
                        packets_processed += 1

                        elapsed_time = time.time() - start_time
                        if elapsed_time >= 60:  # Check if one minute has passed
                            print(f"Processed {packets_processed} packets in one minute.")
                            break
                    else:
                        break
            fp.close()

        return results

    def map_results_to_conversation(self, results, pcap_path):
        counter = 0
        conv_dict = {}
        with open(pcap_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            for packet in rd_ft:
                if counter < len(results):
                    if packet:
                        packet = packet[0].split(',')
                        result = results[counter]
                        conv_number = packet[20]
                        if conv_number not in conv_dict:
                            conv_dict[conv_number] = []
                        conv_dict[conv_number].append(result)
                        counter += 1
                else:
                    break
            fp.close()
        return conv_dict

    def map_results_to_conversation_tuple(self, results, pcap_path):
        counter = 0
        conv_dict = {}
        with open(pcap_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            for packet in rd_ft:
                if counter < len(results):
                    if packet:
                        packet = packet[0].split(',')
                        result = results[counter]
                        conv_number = packet[23]
                        if conv_number not in conv_dict:
                            conv_dict[conv_number] = []
                        conv_dict[conv_number].append({counter: result})
                        counter += 1
                else:
                    break
            fp.close()
        return conv_dict

    def most_significant_packets_sampler(self, day, threshold):
        root_folder = "."
        attack_types_folder = os.path.join(root_folder, "input_data/attack_types")
        pickles_folder = os.path.join(root_folder, "pickles/output_pickles_packet_basis")

        for attack_type in os.listdir(attack_types_folder):
            if attack_type == f"{day}_features.csv" or attack_type == f"{day}_BENIGN.csv" or not (attack_type.startswith(day) and attack_type.endswith(".csv")):
                continue
            attack_type = attack_type.replace(".csv", "")
            attack_type = attack_type.replace(f"{day}_features_", "")

            # Construct the file names for features and pickle file
            feature_file_name = f"{day}_features_{attack_type}.csv"
            pickle_file_name = f"{day.title()}_{attack_type}_results.pkl"
            feature_file_path = os.path.join(attack_types_folder, feature_file_name)
            pickle_file_path = os.path.join(pickles_folder, pickle_file_name)
            print(f'Sampling {attack_type}')
            # Check if the pickle file exists
            if not os.path.exists(pickle_file_path):
                print(f'pickle for {attack_type} not found')
                continue
            # Load the pickle file containing reconstruction errors
            with open(pickle_file_path, 'rb') as pickle_file:
                reconstruction_errors = pickle.load(pickle_file)
            # Load the corresponding feature CSV file
            features_df = pd.read_csv(feature_file_path, header=None)
            # Load the conversation scores
            conv_scores = self.map_results_to_conversation_tuple(reconstruction_errors, f"input_data/attack_types/{day}_{attack_type}.pcap.tsv")
            max_packets = []
            for conv in conv_scores:
                max_dict = max(conv_scores[conv], key=lambda x: list(x.values())[0])
                max_packets.append(max_dict)

            true_positive = []
            for item in max_packets:
                value = list(item.values())[0]  # Extracting the value from the dictionary
                if value > threshold:
                    true_positive.append(item)
                else:
                    true_positive.append(item)
            # Sort list
            sentinel = False
            if 'benign' in attack_type:
                sentinel = True
            # true_positive can indicate a true positive OR a true negative. Same is true for false_...
            true_positive = sorted(true_positive, key=lambda x: list(x.values())[0], reverse=sentinel)
            sorted_keys_tp = [list(d.keys())[0] for d in true_positive]
            # Extract the significant features
            significant_features_tp = features_df.iloc[sorted_keys_tp]
            if len(significant_features_tp) > 0:
                if len(significant_features_tp) > 40:
                    significant_features_tp = significant_features_tp.sample(n=40)
                # Define the output file name
                print(f'writing {attack_type} to file')
                output_file_name = f"{day}_features_{attack_type}_most_significant.csv"
                output_file_path = os.path.join(attack_types_folder, output_file_name)
                # Save the significant features to a new CSV file
                significant_features_tp.to_csv(output_file_path, index=False, header=False)

    def shap_values_builder_from_features(self, test_feature_path, benign_feature_path):
        path = 'pickles/anomDetectorFullDataset.pkl'
        with open(path, 'rb') as f:
            kit = pickle.load(f)

        def callKit(featureList):
            results = []
            for features in featureList:
                results.append(kit.execute(features))
            return np.array(results)

        # Load CSV file since it probably will not be too big
        with open(test_feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            test_features = []
            for feature in rd_ft:
                feature = feature[0].split(',')
                feature = [float(element) for element in feature]
                feature = np.array(feature)
                test_features.append(feature)
            fp.close()
        print('Done building training feature array')

        # Load CSV file since it probably will not be too big
        with open(benign_feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            benign_features = []
            for feature in rd_ft:
                feature = feature[0].split(',')
                feature = [float(element) for element in feature]
                feature = np.array(feature)
                benign_features.append(feature)
            fp.close()

        print("Building SHAP explainer")
        explainer = shap.KernelExplainer(callKit, np.array(benign_features[:40]))
        print("Calculating SHAP values")
        self.shap_values = explainer.shap_values(np.array(test_features[:40]))
        return self.shap_values

    # Calculates SHAP-values for each available attack type in a day of the week; writes results to Excel and pickles results
    def shap_documenter(self, day):
        root_folder = "."
        attack_types_folder = os.path.join(root_folder, "input_data/attack_types")
        self.workbook = openpyxl.load_workbook(f'input_data/template_statistics_file.xlsx')
        for attack_type in os.listdir(attack_types_folder):
            if not (attack_type.startswith(day) and attack_type.endswith("most_significant.csv")):
                continue
            attack_type = attack_type.replace(".csv", "")
            attack_type = attack_type.replace(f"{day}_features_", "")
            # Loop over the different Kitsune configs we are going to make
            shap_values = self.shap_values_builder_from_features(
                f"input_data/attack_types/{day}_features_{attack_type}.csv",
                "input_data/attack_types/monday_features_sample_medium_validate2.csv")
            
            path = f'pickles/output_pickles/{day.title()}_{attack_type}shap_results.pkl'
            with open(path, 'wb') as f:
               pickle.dump(shap_values, f)
            # Could do this with a Regular Expression, but I'm a sane person
            #with open (f'pickles/output_pickles/{day.title()}_{attack_type}shap_results.pkl', 'rb') as f:
            #   self.shap_values = pickle.load(f)
            #self.create_sheet(day, attack_type.replace("most_significant", "").replace("-", "").replace("_", "").replace(" ", ""))
        #excel_file = f"output_data/shap_{day}_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
        #self.workbook.save(excel_file)

    def train_kitsune(self):
        with open(f"input_data/attack_types/monday_features.csv", newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            line_count = sum(1 for row in csv_reader)

        kit = KitNET(self.num_features, max_autoencoder_size=75, FM_grace_period=int(0.05 * line_count),
                     AD_grace_period=line_count, learning_rate=0.001,
                     hidden_ratio=0.25)
        # Load the feature list beforehand to save time

        print('training Kitsune')
        counter = 0
        with open(f"input_data/attack_types/monday_features.csv") as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            train_err = []
            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    if counter % 10000 == 0:
                        print("training: " + str(counter))
                    train_err.append(kit.train(packet))
                    counter += 1
            fp.close()

        median_value = np.median(train_err)
        median_absolute_deviation = np.median([abs(number - median_value) for number in train_err])
        print('done training')
        threshold = median_value + 2 * median_absolute_deviation
        print(f'threshold: {threshold}')
        with open("pickles/anomDetectorFullDataset.pkl", 'wb') as f:
            pickle.dump(kit, f)
